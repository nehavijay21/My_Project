from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils.timezone import now
from .models import Programme, Room, Course, Exam, Timetable, Teacher, DutyAllotment, DutyPreference
from .forms import ProgramForm, RoomForm, CourseForm, ExamForm, TimetableForm, TeacherForm, DutyAllotmentForm, DutyPreferenceForm
from datetime import datetime
from django.utils.dateparse import parse_date 
from django.urls import reverse
import pandas as pd
from django.shortcuts import render, redirect
from .models import ExamAttendance, Student, StudentExam
from django.contrib import messages
from datetime import datetime
import os
from django.conf import settings
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter, range_boundaries




@login_required()
def index(request):
    return render(request, 'index.html')

@login_required()
def manage_programs(request):
    return render(request, 'manage_programs.html')

@login_required()
def manage_rooms(request):
    return render(request, 'manage_rooms.html')

@login_required()
def manage_course(request):
    return render(request, 'manage_course.html')

@login_required()
def manage_exam(request):
    return render(request, 'manage_exam.html')

@login_required()
def manage_timetable(request):
    return render(request, 'manage_timetable.html')

@login_required()
def manage_teacher(request):
    return render(request, 'manage_teacher.html')

@login_required()
def manage_duty(request):
    duties = DutyAllotment.objects.all() 
    return render(request, 'duty_allotment.html', {'duties': duties})

@login_required()
def manage_preference(request):
    return render(request, 'manage_preference.html')

def chief_group_required(user):
    return user.groups.filter(name='Chief').exists()

def teacher_group_required(user):
    return user.groups.filter(name='Teacher').exists()

@login_required
def dashboard_view(request):
    user = request.user
    is_teacher = user.groups.filter(name='teacher').exists()
    is_chief = user.groups.filter(name='chief').exists()
    ongoing_exams = Exam.objects.filter(timetable__date__gte=now().date()).distinct()
    if is_teacher:
        duty_allotments = DutyAllotment.objects.filter(teacher=user.teacher)
        duty_preferences = DutyPreference.objects.filter(teacher=user.teacher)
    else: 
        duty_allotments = DutyAllotment.objects.all()
        duty_preferences = DutyPreference.objects.all()

    context = {
        'is_teacher': is_teacher,
        'is_chief': is_chief,
        'ongoing_exams': ongoing_exams,
        'duty_allotments': duty_allotments,
        'duty_preferences': duty_preferences
    }
    return render(request, 'index.html', context)
    
@login_required()
@user_passes_test(chief_group_required)
def program_list(request):
    programs = Programme.objects.all()
    return render(request, 'program_list.html', {'programs': programs})

@login_required()
@user_passes_test(chief_group_required)
def add_program(request):
    if request.method == 'POST':
        form = ProgramForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('program_list')
    else:
        form = ProgramForm()
    return render(request, 'add_program.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_program(request, pk):
    program = get_object_or_404(Programme, pk=pk)
    if request.method == 'POST':
        form = ProgramForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            return redirect('program_list')
    else:
        form = ProgramForm(instance=program)
    return render(request, 'add_program.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def delete_program(request, pk):
    program = get_object_or_404(Programme, pk=pk)
    if request.method == 'POST':
        program.delete()
        return redirect('program_list')
    return render(request, 'delete_program.html', {'program': program})

@login_required()
@user_passes_test(chief_group_required)
def room_list(request):
    rooms = Room.objects.all()
    return render(request, 'room_list.html', {'rooms': rooms})

@login_required()
@user_passes_test(chief_group_required)
def add_room(request):
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm()
    return render(request, 'add_room.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_room(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm(instance=room)
    return render(request, 'add_room.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def delete_room(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        room.delete()
        return redirect('room_list')
    return render(request, 'delete_room.html', {'room': room})

from django.db.models import IntegerField
from django.db.models.functions import Cast

import re
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Course  

@login_required()
@user_passes_test(chief_group_required)
def course_list(request):
    order = request.GET.get('order', 'asc')  

    def extract_number(course):
        match = re.search(r'\d+', course.course_code)  
        return int(match.group()) if match else float('inf')  

    courses = list(Course.objects.all())  
    courses.sort(key=lambda c: extract_number(c), reverse=(order == 'desc'))

    return render(request, 'course_list.html', {'courses': courses, 'order': order})




@login_required()
@user_passes_test(chief_group_required)
def add_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('course_list')
    else:
        form = CourseForm()
    return render(request, 'add_course.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            return redirect('course_list')
    else:
        form = CourseForm(instance=course)
    return render(request, 'edit_course.html', {'form': form, 'course': course})

@login_required()
@user_passes_test(chief_group_required)
def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        return redirect('course_list')
    return render(request, 'delete_course.html', {'course': course})


from django.urls import reverse


@login_required()
@user_passes_test(chief_group_required)
def exam_list(request):
    exam = Exam.objects.all().order_by('-active', 'sem', 'year')
    return render(request, 'exam_list.html', {'exam': exam})


@login_required()
@user_passes_test(chief_group_required)
def add_exam(request):
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('exam_list')
    else:
        form = ExamForm()
    return render(request, 'add_exam.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            return redirect('exam_list')
    else:
        form = ExamForm(instance=exam)
    return render(request, 'add_exam.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def delete_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        exam.delete()
        return redirect('exam_list')
    return render(request, 'delete_exam.html', {'exam': exam})

@login_required()
@user_passes_test(chief_group_required)
def timetable_list(request):
    timetable= Timetable.objects.all()
    return render(request, 'timetable_list.html', {'timetable': timetable})

@login_required()
@user_passes_test(chief_group_required)
def add_timetable(request):
    ongoing_exams = Exam.objects.filter(active=True)  

    if request.method == 'POST':
        form = TimetableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('timetable_list')
    else:
        form = TimetableForm()

    form.fields['exam'].queryset = ongoing_exams

    return render(request, 'add_timetable.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_timetable(request, pk):
    timetable = get_object_or_404(Timetable, pk=pk)
    ongoing_exams = Exam.objects.filter(active=True)

    if request.method == 'POST':
        form = TimetableForm(request.POST, instance=timetable)
        if form.is_valid():
            form.save()
            return redirect('timetable_list')
    else:
        form = TimetableForm(instance=timetable)

    form.fields['exam'].queryset = ongoing_exams

    return render(request, 'add_timetable.html', {'form': form})


@login_required()
@user_passes_test(chief_group_required)
def delete_timetable(request, pk):
    timetable= get_object_or_404(Timetable, pk=pk)
    if request.method == 'POST':
        timetable.delete()
        return redirect('timetable_list')
    return render(request, 'delete_timetable.html', {'timetable': timetable})


from django.http import JsonResponse
from .models import Course, Exam

def get_courses_by_exam(request):
    exam_id = request.GET.get('exam_id')
    if exam_id:
        try:
            exam = Exam.objects.get(pk=exam_id)
            courses = Course.objects.filter(sem=exam.sem)
            data = [{'id': c.pk, 'name': f'{c.course_code} - {c.course_title}'} for c in courses]
            return JsonResponse({'courses': data})
        except Exam.DoesNotExist:
            return JsonResponse({'error': 'Invalid exam ID'}, status=400)
    return JsonResponse({'error': 'No exam ID provided'}, status=400)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from .models import Teacher
from .forms import TeacherForm, TeacherEditForm  

@login_required()
@user_passes_test(chief_group_required)
def teacher_list(request):
    teachers = Teacher.objects.all()
    return render(request, 'teacher_list.html', {'teachers': teachers})

@login_required()
@user_passes_test(chief_group_required)
def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Teacher added successfully!")
                return redirect('teacher_list')
            except Exception as e:
                messages.error(request, f"Error saving teacher: {e}")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = TeacherForm()
    
    return render(request, 'add_teacher.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def edit_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    user = teacher.user

    if request.method == "POST":
        form = TeacherEditForm(request.POST, instance=teacher)  
        if form.is_valid():
            form.save()
            messages.success(request, "Teacher updated successfully!")
            return redirect('teacher_list')
    else:
        form = TeacherEditForm(instance=teacher)

    return render(request, 'add_teacher.html', {'form': form, 'edit_mode': True})

@login_required()
@user_passes_test(chief_group_required)
def delete_teacher(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.user.delete()
        return redirect('teacher_list')
    return render(request, 'delete_teacher.html', {'teacher': teacher})

@login_required()
@user_passes_test(chief_group_required)
def change_teacher_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = PasswordChangeForm(user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password updated successfully.')
            return redirect('teacher_list')
    else:
        form = PasswordChangeForm(user)

    full_name = f"{user.first_name} {user.last_name}"
    return render(request, 'change_password.html', {'form': form, 'full_name': full_name})


from django.contrib.auth.forms import SetPasswordForm

@login_required()
@user_passes_test(chief_group_required)
def reset_teacher_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Password reset successfully for {user.username}.')
            return redirect('teacher_list')
    else:
        form = SetPasswordForm(user)

    full_name = f"{user.first_name} {user.last_name}"
    return render(request, 'reset_password.html', {'form': form, 'full_name': full_name})


@login_required()
@user_passes_test(chief_group_required)
def duty_allotment(request):
    exam_dates = Timetable.objects.all().order_by('date')

    duty_counts = {date.date: DutyAllotment.objects.filter(date=date.date).count() for date in exam_dates}

    return render(request, 'duty_allotment.html', {'exam_dates': exam_dates, 'duty_counts': duty_counts})

@login_required()
@user_passes_test(chief_group_required)
def manage_duty_allotments(request):
    selected_date = request.GET.get('date')  
    if selected_date:
        duties = DutyAllotment.objects.filter(date=selected_date)  
    else:
        duties = DutyAllotment.objects.all()  

    return render(request, 'manage_duty_allotments.html', {'duties': duties, 'selected_date': selected_date})


@login_required()
@user_passes_test(chief_group_required)
def duty_list(request):
    date_str = request.GET.get('date')  
    formatted_date = None  

    if date_str:
        formatted_date = parse_date(date_str)  
        if not formatted_date:
            try:
                formatted_date = datetime.strptime(date_str, "%b. %d, %Y").date()  # Handles 'Feb. 19, 2025' format
            except ValueError:
                formatted_date = None 
    duties = DutyAllotment.objects.filter(date=formatted_date, teacher__isnull=False) if formatted_date else []

    return render(request, 'duty_list.html', {'duties': duties, 'selected_date': formatted_date})


from django.utils.dateparse import parse_date
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Teacher, DutyAllotment
from .forms import DutyAllotmentForm

@login_required()
@user_passes_test(chief_group_required)
def add_duty(request):
    selected_date = request.GET.get('date', None)
    
    print("Received date from URL:", selected_date)

    if not selected_date:
        messages.error(request, "No date provided. Please select a valid date.")
        return redirect('duty_allotment')
    formatted_date = parse_date(selected_date)

    if not formatted_date:
        messages.error(request, f"Invalid date format received: {selected_date}")
        return redirect('duty_allotment')

    print("Parsed date:", formatted_date)
    preferred_teachers = Teacher.objects.filter(duty_preferences__pref_date=formatted_date).distinct()

    for teacher in preferred_teachers:
        teacher.duty_counts = DutyAllotment.objects.filter(teacher=teacher, date=formatted_date).count()

    if request.method == 'POST':
        form = DutyAllotmentForm(request.POST)
        if form.is_valid():
            duty = form.save(commit=False)
            duty.date = formatted_date 
            duty.save()
            messages.success(request, "Duty allotted successfully!")
            return redirect(reverse('duty_list') + f"?date={selected_date}")
        else:
            messages.error(request, "Form is invalid. Please check the entered data.")
    else:
        form = DutyAllotmentForm(initial={'date': formatted_date})

    return render(request, 'add_duty.html', {
        'form': form,
        'selected_date': formatted_date.strftime('%Y-%m-%d'),
        'preferred_teachers': preferred_teachers,
    })


    
@login_required()
@user_passes_test(chief_group_required)
def edit_duty(request, pk):
    duty = get_object_or_404(DutyAllotment, pk=pk)
    if request.method == 'POST':
        form = DutyAllotmentForm(request.POST, instance=duty)
        if form.is_valid():
            form.save()
            return redirect('duty_list')
    else:
        form = DutyAllotmentForm(instance=duty)
    return render(request, 'add_duty.html', {'form': form})

@login_required()
@user_passes_test(chief_group_required)
def delete_duty(request, pk):
    duty = get_object_or_404(DutyAllotment, pk=pk)
    selected_date = duty.date.strftime("%Y-%m-%d") 

    if request.method == 'POST':
        duty.delete()
        messages.success(request, "Duty deleted successfully!")
        return redirect(f"{reverse('duty_list')}?date={selected_date}")

    return render(request, 'delete_duty.html', {'duty': duty})

@login_required()
def preference_list(request):
    teacher=Teacher.objects.get(user=request.user)
    preferences = DutyPreference.objects.filter(teacher=teacher)
    return render(request, 'preference_list.html', {'preferences': preferences})

@login_required()
@user_passes_test(teacher_group_required)
def add_preference(request):
    teacher = Teacher.objects.get(user=request.user)

    if request.method == 'POST':
        form = DutyPreferenceForm(request.POST, user=teacher)
        if form.is_valid():
            selected_dates = form.cleaned_data['pref_dates']
            for date in selected_dates:
                DutyPreference.objects.create(teacher=teacher, pref_date=date)
            messages.success(request, "Preferences saved successfully.")
            return redirect('preference_list')
    else:
        form = DutyPreferenceForm(user=teacher)

    return render(request, 'add_preference.html', {'form': form})

@login_required()
@user_passes_test(teacher_group_required)
def edit_preference(request, pk):
    preference = get_object_or_404(DutyPreference, pk=pk)
    if request.method == 'POST':
        form = DutyPreferenceForm(request.POST, instance=preference)
        if form.is_valid():
            form.save()
            return redirect('preference_list')
    else:
        form = DutyPreferenceForm(instance=preference)
    return render(request, 'add_preference.html', {'form': form})

@login_required()
@user_passes_test(teacher_group_required)
def delete_preference(request, pk):
    preference = get_object_or_404(DutyPreference, pk=pk)
    if request.method == 'POST':
        preference.delete()
        return redirect('preference_list')
    return render(request, 'delete_preference.html', {'preference': preference})

@login_required
@user_passes_test(teacher_group_required)
def duty_history(request):
    user = request.user
    if hasattr(user, 'teacher'):
        teacher = user.teacher
        duty_records = DutyAllotment.objects.filter(teacher=teacher)
    else:
        duty_records = []

    return render(request, 'duty_history.html', {'duty_records': duty_records})

from django.shortcuts import render
from .models import DutyAllotment
from .forms import DateFilterForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count

@login_required
@user_passes_test(chief_group_required)
def chief_duty_history(request):
    form = DateFilterForm(request.GET or None)

    duty_records = DutyAllotment.objects.all()

    if form.is_valid():
        from_date = form.cleaned_data.get('from_date')
        to_date = form.cleaned_data.get('to_date')

        if from_date and to_date:
            duty_records = duty_records.filter(date__range=(from_date, to_date))

    duty_summary = (
        duty_records.values('teacher__user__username')
        .annotate(total_duties=Count('id'))
        .order_by('teacher__user__username')
    )

    return render(request, 'chief_duty_history.html', {
        'form': form,
        'duty_summary': duty_summary
    })



@login_required()
@user_passes_test(chief_group_required)
def upload_nominal_roll(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect('setup_nominal_roll')

        try:
            df = pd.read_excel(excel_file)
            if 'Register Number' not in df.columns:
                messages.error(request, "Missing 'Register Number' column.")
                return redirect('setup_nominal_roll')

            for index, row in df.iterrows():
                register_number = row['Register Number']
                scribe_number = row.get('Scribe', '')
                scribe_number = scribe_number if pd.notna(scribe_number) and str(scribe_number).strip() else None

                student, _ = Student.objects.get_or_create(
                    register_number=register_number,
                    defaults={'scribe_number': scribe_number}
                )

                if student.scribe_number != scribe_number and scribe_number:
                    student.scribe_number = scribe_number
                    student.save()

                for col in df.columns:
                    if col not in ['Register Number', 'Scribe']:
                        date = pd.to_datetime(col, dayfirst=True).date()
                        cell_value = row[col]
                        if pd.notna(cell_value) and '-' in str(cell_value):
                            try:
                                course_code, course_title = map(str.strip, str(cell_value).split('-', 1))

                                exam_attendance, _ = ExamAttendance.objects.get_or_create(
                                    date=date,
                                    course_code=course_code,
                                    course_title=course_title
                                )

                                StudentExam.objects.get_or_create(
                                    exam_attendance=exam_attendance,
                                    student=student
                                )
                            except Exception as e:
                                print(f"⚠️ Error parsing course info in row {index}: {e}")
                                continue

            messages.success(request, "Nominal Roll uploaded successfully!")
        except Exception as e:
            messages.error(request, f"Error uploading file: {str(e)}")

        return redirect('setup_nominal_roll')
    summary = {}
    for attendance in ExamAttendance.objects.all().order_by('date'):
        student_exams = StudentExam.objects.filter(exam_attendance=attendance)
        total_students = student_exams.count()
        absent = student_exams.filter(is_absent=True).count()
        present = total_students - absent

        summary.setdefault(attendance.date, []).append({
            'course_code': attendance.course_code,
            'course_title': attendance.course_title,
            'total': total_students,
            'present': present,
            'absent': absent,
        })

    context = {
        'summary': summary,
    }
    return render(request, 'setup_nominal_roll.html', context)



@login_required()
@user_passes_test(chief_group_required)

def delete_nominal_roll(request):
    StudentExam.objects.all().delete()
    ExamAttendance.objects.all().delete()
    Student.objects.all().delete()

    messages.success(request, "All nominal roll entries deleted.")
    return redirect('setup_nominal_roll')

from urllib.parse import unquote
@login_required
@user_passes_test(chief_group_required)
def mark_attendance(request, date, course_code):
    try:
        exam_attendance = ExamAttendance.objects.get(date=date, course_code=course_code)
    except ExamAttendance.DoesNotExist:
        return HttpResponseNotFound("Exam not found.")
    except ExamAttendance.MultipleObjectsReturned:
        return HttpResponseServerError("Multiple exams found for same course code and date.")

    student_exams = StudentExam.objects.filter(exam_attendance=exam_attendance)

    if request.method == 'POST':
        if 'clear' in request.POST:
            student_exams.update(is_absent=False)
        else:
            absent_ids = request.POST.getlist('absent')
            for se in student_exams:
                se.is_absent = se.student.register_number in absent_ids
                se.save()
        return redirect('setup_nominal_roll')

    total = student_exams.count()
    absent = student_exams.filter(is_absent=True).count()
    present = total - absent

    formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d-%m-%Y")

    context = {
    'date': formatted_date,
    'course_code': course_code,
    'course_title': exam_attendance.course_title,  # 👈 ADD THIS
    'student_exams': student_exams,
    'total': total,
    'present': present,
    'absent': absent,
    }
    return render(request, 'mark_attendance.html', context)

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
from datetime import date
from django.conf import settings
from django.http import HttpResponse
from .models import ExamAttendance, StudentExam  # adjust import if needed
import subprocess
from django.db.models.functions import TruncDate


def create_seating_excel(date_of_exam, course_data_list):
    wb = Workbook()
    first_course = course_data_list[0]
    course_code = first_course['course_code']
    course_title = first_course['course_title']
    absentees = first_course.get('absentees', [])
    exam_date = first_course.get('date_of_exam') 

    sheet_title = f"{course_code} - {course_title}"
    if len(sheet_title) > 31:
        sheet_title = sheet_title[:31]

    ws = wb.active
    ws.title = sheet_title

    # ✅ Calculate present_count before calling setup_sheet
    total_students = StudentExam.objects.filter(
        exam_attendance__course_code=course_code,
        exam_attendance__date=exam_date
    ).count()
    present_count = total_students - len(absentees)


    def setup_sheet(ws, course_code, course_title, absentees, date_of_exam, present_count):
        ws.row_dimensions[8].height = 28.35
        ws.row_dimensions[16].height = 28.35

        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True

        headers = [
            ('A1:I1', "KANNUR UNIVERSITY"),
            ('A2:I2', "STATEMENT TO ACCOMPANY ANSWER PAPER PACKETS SENT TO "),
            ('A3:I3', "THE CONTROLLER OF EXAMINATIONS BY THE CHIEF SUPERINTENDENT")
        ]

        for cell_range, text in headers:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = text
            cell.font = Font(bold=True, name="Ubuntu", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row_start = 4
        merge_sections = [
            (f'A{row_start}:I{row_start}', "Centre of Examination:Nehru Arts and Science College Kanhangad"),
            (f'A{row_start+1}:I{row_start+1}', "Name of Examination:"),
            (f'A{row_start+2}:I{row_start+2}', "No. of Packets:")
        ]

        ws.merge_cells('A6:I7')
        ws['A6'].value = ""
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A6'].font = Font(name="Ubuntu", size=11)

        for cell_range, text in merge_sections:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = text
            cell.font = Font(name="Ubuntu", size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        subject_header_row = row_start + 4
        ws.cell(row=subject_header_row, column=1).value = "Date"
        ws.cell(row=subject_header_row, column=2).value = "Hours"
        ws.merge_cells(start_row=subject_header_row, start_column=3, end_row=subject_header_row, end_column=5)
        ws.cell(row=subject_header_row, column=3).value = "Subject"
        ws.merge_cells(start_row=subject_header_row, start_column=6, end_row=subject_header_row, end_column=7)
        ws.cell(row=subject_header_row, column=6).value = "Code No. of Question Paper"
        ws.merge_cells(start_row=subject_header_row, start_column=8, end_row=subject_header_row, end_column=9)
        ws.cell(row=subject_header_row, column=8).value = "Total No. of Answer Books"

        for col in range(1, 10):
            cell = ws.cell(row=subject_header_row, column=col)
            cell.font = Font(bold=True, name="Ubuntu", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        ws.merge_cells('A9:A13')
        ws.merge_cells('B9:B13')
        ws.merge_cells('C9:E13')
        subject_cell = ws['C9']
        subject_cell.value = f"{course_code} - {course_title}"
        subject_cell.font = Font(name="Ubuntu", size=10)
        subject_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        subject_cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.merge_cells('F9:G13')
        ws.merge_cells('H9:I13')
        ws['H9'].value = str(present_count)
        date_cell = ws['A9']
        date_cell.value = date_of_exam.strftime('%d-%m-%Y')
        date_cell.font = Font(name="Ubuntu", size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        date_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Fill in Hours (hardcoded as "3 hrs")
        hour_cell = ws['B9']
        hour_cell.value = "3 hrs"
        hour_cell.font = Font(name="Ubuntu", size=10)
        hour_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hour_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


        for cell_range in ['A9', 'B9', 'C9', 'F9', 'H9']:
            cell = ws[cell_range]
            cell.font = Font( name="Ubuntu", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        for r in range(subject_header_row + 1, subject_header_row + 6):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        books_header_row = subject_header_row + 7
        ws.merge_cells(f'A{books_header_row}:I{books_header_row}')
        ws[f'A{books_header_row}'].value = "REGISTER NUMBER OF BOOKS"
        ws[f'A{books_header_row}'].font = Font(bold=True, name="Ubuntu", size=10)
        ws[f'A{books_header_row}'].alignment = Alignment(horizontal='center', wrap_text=True)

        book_columns = ["From", "To", "No. of Books"] * 3
        for i, header in enumerate(book_columns):
            col = i + 1
            cell = ws.cell(row=books_header_row + 1, column=col)
            cell.value = header
            cell.font = Font(bold=True, name="Ubuntu", size=10)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1, 10):
            ws.merge_cells(start_row=17, start_column=col, end_row=20, end_column=col)
            for row in range(17, 21):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Section 3: REGISTER NUMBER OF ABSENTEES
        absentee_header_row = books_header_row + 7
        ws.merge_cells(f'A{absentee_header_row}:I{absentee_header_row}')
        cell = ws[f'A{absentee_header_row}']
        cell.value = "REGISTER NUMBER OF ABSENTEES"
        cell.font = Font(bold=True, name="Ubuntu", size=10)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Start from the next row after the header (row 23)
        absentee_start_row = absentee_header_row + 1

        # Split absentee list into 5 rows, and each row should hold a maximum of 5 absentees
        max_per_row = 5

        # Loop over absentees and fill them in rows, 5 per row
        for i in range(5):
            line_row = absentee_start_row + i
            ws.merge_cells(start_row=line_row, start_column=1, end_row=line_row, end_column=9)
            
            # Slice the absentee list to get the max 5 absentees for this row
            chunk = absentees[i * max_per_row:(i + 1) * max_per_row]

            # Set the cell value with the absentee data (comma-separated)
            cell = ws.cell(row=line_row, column=1)
            if chunk:  # If there are absentees in the chunk
                cell.value = ", ".join(chunk)  # Comma-separated absentees
            else:
                cell.value = ""  # Leave the cell blank if no absentees
            
            # Set cell style
            cell.font = Font(name="Ubuntu", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Add borders (only from row 24 to 27, not 23)
            if i >= 1:  # Starting from row 24 onwards
                cell.border = Border(
                    top=Side(style='thin'),  # Top border for row 24 and onwards
                    bottom=Side(style='thin')  # Bottom border for all rows 24 to 27
                )

            # Set row height for compact size (~0.68 cm)
            ws.row_dimensions[line_row].height = 19  # Adjust row height to compact size


        footer_row = absentee_header_row + 7
        ws.merge_cells(f'A{footer_row}:I{footer_row}')
        ws[f'A{footer_row}'].value = "Station:Kanhangad"
        ws[f'A{footer_row}'].alignment = Alignment(horizontal='left')

        footer_next_row = footer_row + 1
        ws.merge_cells(f'A{footer_next_row}:D{footer_next_row}')
        ws[f'A{footer_next_row}'].value = f"Date: {date_of_exam.strftime('%d-%m-%Y')}"
        ws[f'A{footer_next_row}'].alignment = Alignment(horizontal='left')
        ws.merge_cells(f'F{footer_next_row}:I{footer_next_row}')
        ws[f'F{footer_next_row}'].value = "Chief Superintendent"
        ws[f'F{footer_next_row}'].font = Font(bold=True, name="Ubuntu", size=10)
        ws[f'F{footer_next_row}'].alignment = Alignment(horizontal='right')

        note_row = footer_row + 3
        ws.merge_cells(f'A{note_row}:I{note_row}')
        ws[f'A{note_row}'].value = (
            "N.B.– This statement should accompany the answer papers for each subject and should be "
            "carefully verified with the answer books by the Chief Superintendent before despatch."
        )
        ws[f'A{note_row}'].alignment = Alignment(wrap_text=True, horizontal='left')
        ws.row_dimensions[note_row].height = 30
    setup_sheet(ws, course_code, course_title, absentees, first_course['date_of_exam'], present_count)


    for course in course_data_list[1:]:
        course_code = course['course_code']
        course_title = course['course_title']
        absentees = course.get('absentees', [])
        sheet_title = f"{course_code} - {course_title}"
        if len(sheet_title) > 31:
            sheet_title = sheet_title[:31]
        ws = wb.create_sheet(title=sheet_title)
        total_students = len(StudentExam.objects.filter(
        exam_attendance__course_code=course_code,
        exam_attendance__date=course['date_of_exam']
    ))
        present_count = total_students - len(absentees)

    setup_sheet(ws, course_code, course_title, absentees, course['date_of_exam'], present_count)


    safe_date = date_of_exam.strftime("%Y-%m-%d")
    file_name = f"Seating_Layout_{safe_date}.xlsx"
    folder_path = os.path.join(settings.MEDIA_ROOT, 'seating_files')
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, file_name)
    wb.save(file_path)
    print(f"File saved as {file_path}")

def generate_excel(request):
    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'error': 'No date provided'}, status=400)

    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%d-%m-%Y").date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    exams = ExamAttendance.objects.filter(date=date_obj).order_by('course_code')
    if not exams.exists():
        return JsonResponse({'error': 'No exams found for this date'}, status=404)

    course_data_list = []
    for exam in exams:
        absentees_qs = StudentExam.objects.filter(exam_attendance=exam, is_absent=True)
        absentees = [se.student.register_number for se in absentees_qs]
        course_data_list.append({
            'course_code': exam.course_code,
            'course_title': exam.course_title,
            'absentees': absentees,
            'date_of_exam': date_obj,
        })

    if not course_data_list:
        return JsonResponse({'error': 'No data for this date'}, status=404)

    create_seating_excel(date_obj, course_data_list)

    safe_date = date_obj.strftime("%Y-%m-%d")
    excel_path = os.path.join(settings.MEDIA_ROOT, 'seating_files', f"Seating_Layout_{safe_date}.xlsx")

    try:
        convert_excel_to_pdf(excel_path)
    except Exception as e:
        print(f"PDF conversion failed for {safe_date}: {e}")
        return JsonResponse({'error': 'PDF conversion failed'}, status=500)

    pdf_url = os.path.join(settings.MEDIA_URL, 'seating_files', f"Seating_Layout_{safe_date}.pdf")
    return JsonResponse({'pdf_url': pdf_url})

def convert_excel_to_pdf(input_file):
    folder_path = os.path.join(settings.MEDIA_ROOT, 'seating_files')
    command = [
        "soffice", "--headless", "--convert-to", "pdf",
        "--outdir", folder_path, input_file
    ]
    subprocess.run(command, check=True)

from django.http import JsonResponse
from .models import ExamAttendance

def get_exam_dates(request):
    dates = ExamAttendance.objects.values_list('date', flat=True).distinct()
    formatted_dates = sorted(set(d.strftime('%d-%m-%Y') for d in dates if d))
    return JsonResponse({'dates': formatted_dates})